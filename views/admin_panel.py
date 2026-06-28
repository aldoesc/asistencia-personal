import flet as ft
import os
import shutil
from datetime import date, timedelta
from backend.database import SessionLocal
from backend.crud import (
    get_sucursales, crear_sucursal, eliminar_sucursal,
    get_empleados, crear_empleado, actualizar_empleado, desactivar_empleado,
    DNIExistenteError,
    set_horario_base, get_horarios_empleado,
    get_registros_empleado, get_empleado_por_codigo,
    cambiar_password_admin,
)
from backend.logic import get_turno_nombre
from backend.models import TipoTurno, DiaSemana
from utils.qr_utils import generar_qr, get_qr_path
from backend.database import DB_PATH


# ── Exportación ───────────────────────────────────────────────────────────────

def _exportar_excel(filas_salud, filas_asistencia=None):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "Salud de Responsabilidad"
    headers = ["Empleado","Cargo","A tiempo","T.Baja","T.Media","T.Grave","Ausente","Salud %","Nivel"]
    header_fill = PatternFill("solid", fgColor="1c2128")
    header_font = Font(bold=True, color="AAAAAA")
    thin = Side(style="thin", color="333333")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal="center"); cell.border = border
        ws.column_dimensions[cell.column_letter].width = max(14, len(h) + 4)
    for fila in filas_salud:
        ws.append(fila)
    if filas_asistencia:
        ws2 = wb.create_sheet("Asistencia")
        for col, h in enumerate(["Fecha","Turno","Ingreso","Salida","Estado","Justificado"], 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill = header_fill; cell.font = header_font; cell.border = border
        for fila in filas_asistencia:
            ws2.append(fila)
    escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
    os.makedirs(escritorio, exist_ok=True)
    ruta = os.path.join(escritorio, f"reporte_{date.today().isoformat()}.xlsx")
    wb.save(ruta)
    return ruta


def _exportar_pdf(filas_salud, filas_asistencia=None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
    os.makedirs(escritorio, exist_ok=True)
    ruta = os.path.join(escritorio, f"reporte_{date.today().isoformat()}.pdf")
    doc = SimpleDocTemplate(ruta, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Reporte de Salud de Responsabilidad", styles["Title"]),
        Paragraph(f"Generado: {date.today().strftime('%d/%m/%Y')}", styles["Normal"]),
        Spacer(1, 0.4*cm),
    ]
    headers = ["Empleado","Cargo","A tiempo","T.Baja","T.Media","T.Grave","Ausente","Salud %","Nivel"]
    t = Table([headers] + filas_salud, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1c2128")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),9),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.HexColor("#0f1117"),colors.HexColor("#161b22")]),
        ("TEXTCOLOR",(0,1),(-1,-1),colors.HexColor("#cccccc")),
        ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#21262d")),
        ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
    ]))
    story.append(t)
    doc.build(story)
    return ruta


# ── Panel principal ───────────────────────────────────────────────────────────

class AdminPanel:
    def __init__(self, page: ft.Page):
        self.page = page
        self.dialog = None
        self._filas_asistencia_export = []

    def open(self):
        tabs = ft.Tabs(
            selected_index=0,
            indicator_color=ft.Colors.BLUE_400,
            label_color=ft.Colors.BLUE_300,
            unselected_label_color=ft.Colors.GREY_500,
            tabs=[
                ft.Tab(text="🏪 Tiendas",    content=self._tab_tiendas()),
                ft.Tab(text="👥 Personal",   content=self._tab_personal()),
                ft.Tab(text="📅 Horarios",   content=self._tab_horarios()),
                ft.Tab(text="📋 Asistencia", content=self._tab_asistencia()),
                ft.Tab(text="📊 Reportes",   content=self._tab_reportes()),
                ft.Tab(text="🔒 Seguridad",  content=self._tab_seguridad()),
            ],
            expand=True,
        )
        self.dialog = ft.AlertDialog(
            modal=True,
            title=ft.Row([
                ft.Icon(ft.Icons.ADMIN_PANEL_SETTINGS_ROUNDED,
                        color=ft.Colors.BLUE_400, size=22),
                ft.Text("Panel de Administración", size=18,
                        weight=ft.FontWeight.BOLD),
            ], spacing=10),
            content=ft.Container(content=tabs, width=1020, height=700,
                                 padding=ft.padding.only(top=10)),
            actions=[ft.TextButton("Cerrar", on_click=lambda _: self.close(),
                                   style=ft.ButtonStyle(color=ft.Colors.GREY_400))],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        self.page.open(self.dialog)

    def close(self):
        if self.dialog:
            self.page.close(self.dialog)

    def _snack(self, msg: str, error=False):
        self.page.open(ft.SnackBar(
            content=ft.Text(msg),
            bgcolor=ft.Colors.RED_900 if error else ft.Colors.GREY_800,
        ))

    # ── TAB TIENDAS ───────────────────────────────────────────────────────────
    def _tab_tiendas(self):
        nombre_field = ft.TextField(
            label="Nombre de la tienda", width=320, border_radius=8,
            hint_text="Ej: Tienda Miraflores",
            border_color=ft.Colors.BLUE_700,
            focused_border_color=ft.Colors.BLUE_400,
            prefix_icon=ft.Icons.STORE_OUTLINED,
        )
        lista = ft.Column(spacing=8, scroll=ft.ScrollMode.AUTO)

        def cargar():
            lista.controls.clear()
            with SessionLocal() as db:
                for s in get_sucursales(db):
                    sid, snom = s.id, s.nombre
                    lista.controls.append(
                        ft.Container(
                            content=ft.Row([
                                ft.Container(
                                    content=ft.Icon(ft.Icons.STORE,
                                                    color=ft.Colors.BLUE_400, size=18),
                                    width=36, height=36,
                                    bgcolor=ft.Colors.with_opacity(0.1, ft.Colors.BLUE_400),
                                    border_radius=8, alignment=ft.alignment.center,
                                ),
                                ft.Text(snom, size=14, color=ft.Colors.WHITE,
                                        expand=True, weight=ft.FontWeight.W_500),
                                ft.IconButton(
                                    icon=ft.Icons.DELETE_OUTLINE,
                                    icon_color=ft.Colors.RED_400, icon_size=18,
                                    tooltip="Eliminar tienda",
                                    on_click=lambda _, i=sid: _eliminar(i),
                                ),
                            ], vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=12),
                            padding=ft.padding.symmetric(horizontal=16, vertical=10),
                            bgcolor="#1c2128", border_radius=10,
                            border=ft.border.all(1, "#21262d"),
                        )
                    )
            self.page.update()

        def _agregar(e):
            nombre = nombre_field.value.strip()
            if not nombre:
                self._snack("⚠️ Escribe el nombre de la tienda", error=True); return
            with SessionLocal() as db:
                crear_sucursal(db, nombre)
            nombre_field.value = ""
            cargar()
            self._snack(f"✅ Tienda '{nombre}' creada")

        def _eliminar(sid):
            with SessionLocal() as db:
                eliminar_sucursal(db, sid)
            cargar()
            self._snack("Tienda eliminada")

        cargar()
        return ft.Column([
            ft.Text("TIENDAS / SUCURSALES", size=13,
                    weight=ft.FontWeight.BOLD, color=ft.Colors.GREY_400),
            ft.Text("Agrega o elimina las tiendas disponibles al registrar personal.",
                    size=12, color=ft.Colors.GREY_500),
            ft.Container(height=4),
            ft.Row([
                nombre_field,
                ft.ElevatedButton("➕ Agregar", on_click=_agregar,
                                  bgcolor=ft.Colors.BLUE_700, color=ft.Colors.WHITE),
            ], spacing=12, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            ft.Divider(color="#21262d", height=20),
            ft.Text("TIENDAS REGISTRADAS", size=13,
                    weight=ft.FontWeight.BOLD, color=ft.Colors.GREY_400),
            lista,
        ], spacing=12, scroll=ft.ScrollMode.AUTO)

    # ── TAB PERSONAL ─────────────────────────────────────────────────────────
    def _tab_personal(self):
        # ── Campos de creación ──
        f_nombre   = ft.TextField(label="Nombre",  width=220, border_radius=8,
                                  border_color=ft.Colors.BLUE_700,
                                  focused_border_color=ft.Colors.BLUE_400)
        f_apellido = ft.TextField(label="Apellido",width=220, border_radius=8,
                                  border_color=ft.Colors.BLUE_700,
                                  focused_border_color=ft.Colors.BLUE_400)
        f_dni      = ft.TextField(label="DNI",     width=160, border_radius=8,
                                  border_color=ft.Colors.BLUE_700,
                                  focused_border_color=ft.Colors.BLUE_400)
        f_cargo    = ft.TextField(label="Cargo",   width=220, border_radius=8,
                                  border_color=ft.Colors.BLUE_700,
                                  focused_border_color=ft.Colors.BLUE_400)
        f_sucursal = ft.Dropdown(label="Tienda", width=200, border_radius=8)

        def _recargar_sucursales(drop):
            with SessionLocal() as db:
                sucs = get_sucursales(db)
            drop.options = [ft.dropdown.Option(s.nombre) for s in sucs]
            drop.value = sucs[0].nombre if sucs else None

        _recargar_sucursales(f_sucursal)

        # ── Tabla de empleados ──
        tabla = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Código",  color=ft.Colors.GREY_400, size=12)),
                ft.DataColumn(ft.Text("Nombre",  color=ft.Colors.GREY_400, size=12)),
                ft.DataColumn(ft.Text("DNI",     color=ft.Colors.GREY_400, size=12)),
                ft.DataColumn(ft.Text("Cargo",   color=ft.Colors.GREY_400, size=12)),
                ft.DataColumn(ft.Text("Tienda",  color=ft.Colors.GREY_400, size=12)),
                ft.DataColumn(ft.Text("Acciones",color=ft.Colors.GREY_400, size=12)),
            ],
            rows=[],
            heading_row_color="#1c2128",
            border=ft.border.all(1, "#21262d"),
            border_radius=10,
            horizontal_lines=ft.BorderSide(1, "#21262d"),
            column_spacing=16,
        )

        def cargar_empleados():
            tabla.rows.clear()
            with SessionLocal() as db:
                for emp in get_empleados(db):
                    eid       = emp.id
                    enombre   = emp.nombre
                    eapellido = emp.apellido
                    edni      = emp.dni
                    ecargo    = emp.cargo
                    esuc      = emp.sucursal.nombre if emp.sucursal else "—"
                    ecodigo   = emp.codigo_unico

                    tabla.rows.append(ft.DataRow(cells=[
                        ft.DataCell(ft.Text(ecodigo, color=ft.Colors.BLUE_300, size=13)),
                        ft.DataCell(ft.Text(f"{enombre} {eapellido}",
                                            weight=ft.FontWeight.W_500)),
                        ft.DataCell(ft.Text(edni, color=ft.Colors.GREY_400)),
                        ft.DataCell(ft.Text(ecargo, color=ft.Colors.GREY_400)),
                        ft.DataCell(ft.Text(esuc,   color=ft.Colors.GREY_400, size=13)),
                        ft.DataCell(ft.Row([
                            # Botón QR
                            ft.IconButton(
                                icon=ft.Icons.QR_CODE_2,
                                icon_color=ft.Colors.BLUE_400,
                                icon_size=20, tooltip="Ver / Guardar QR",
                                on_click=lambda _, c=ecodigo, n=f"{enombre} {eapellido}":
                                    self._mostrar_qr(c, n),
                            ),
                            # Botón editar
                            ft.IconButton(
                                icon=ft.Icons.EDIT_OUTLINED,
                                icon_color=ft.Colors.AMBER_400,
                                icon_size=20, tooltip="Editar empleado",
                                on_click=lambda _, i=eid, no=enombre, ap=eapellido,
                                                d=edni, ca=ecargo, su=esuc:
                                    self._editar_empleado(
                                        i, no, ap, d, ca, su, cargar_empleados),
                            ),
                            # Botón desactivar
                            ft.IconButton(
                                icon=ft.Icons.PERSON_REMOVE_OUTLINED,
                                icon_color=ft.Colors.RED_400,
                                icon_size=20, tooltip="Desactivar empleado",
                                on_click=lambda _, i=eid: _desactivar(i),
                            ),
                        ], spacing=0)),
                    ]))
            self.page.update()

        def _desactivar(eid):
            with SessionLocal() as db:
                desactivar_empleado(db, eid)
            cargar_empleados()
            self._snack("Empleado desactivado")

        def _crear(e):
            if not all([f_nombre.value, f_apellido.value, f_dni.value, f_cargo.value]):
                self._snack("⚠️ Completa todos los campos", error=True); return
            if not f_sucursal.value:
                self._snack("⚠️ Selecciona una tienda", error=True); return
            try:
                with SessionLocal() as db:
                    emp = crear_empleado(db, f_nombre.value, f_apellido.value,
                                         f_dni.value, f_cargo.value, f_sucursal.value)
                    generar_qr(emp.codigo_unico, emp.nombre_completo)
                cargar_empleados()
                for field in [f_nombre, f_apellido, f_dni, f_cargo]:
                    field.value = ""
                self._snack(f"✅ {emp.codigo_unico} creado · QR generado")
            except DNIExistenteError as ex:
                self._snack(f"⚠️ {ex}", error=True)
            except Exception as ex:
                self._snack(f"❌ Error inesperado: {ex}", error=True)

        cargar_empleados()

        return ft.Column([
            ft.Text("NUEVO EMPLEADO", size=13,
                    weight=ft.FontWeight.BOLD, color=ft.Colors.GREY_400),
            ft.Row([f_nombre, f_apellido], spacing=10),
            ft.Row([f_dni, f_cargo, f_sucursal], spacing=10),
            ft.ElevatedButton("➕ Crear empleado", on_click=_crear,
                              bgcolor=ft.Colors.BLUE_700, color=ft.Colors.WHITE),
            ft.Divider(color="#21262d", height=20),
            ft.Text("EMPLEADOS ACTIVOS", size=13,
                    weight=ft.FontWeight.BOLD, color=ft.Colors.GREY_400),
            ft.Container(content=tabla, border_radius=10,
                         clip_behavior=ft.ClipBehavior.ANTI_ALIAS),
        ], spacing=14, scroll=ft.ScrollMode.AUTO)

    # ── Diálogo editar empleado ───────────────────────────────────────────────
    def _editar_empleado(self, emp_id, nombre, apellido, dni, cargo, sucursal, on_done):
        e_nombre   = ft.TextField(label="Nombre",  value=nombre,   width=240,
                                  border_radius=8, border_color=ft.Colors.BLUE_700,
                                  focused_border_color=ft.Colors.BLUE_400)
        e_apellido = ft.TextField(label="Apellido",value=apellido, width=240,
                                  border_radius=8, border_color=ft.Colors.BLUE_700,
                                  focused_border_color=ft.Colors.BLUE_400)
        e_dni      = ft.TextField(label="DNI",     value=dni,      width=180,
                                  border_radius=8, border_color=ft.Colors.BLUE_700,
                                  focused_border_color=ft.Colors.BLUE_400)
        e_cargo    = ft.TextField(label="Cargo",   value=cargo,    width=240,
                                  border_radius=8, border_color=ft.Colors.BLUE_700,
                                  focused_border_color=ft.Colors.BLUE_400)
        e_sucursal = ft.Dropdown(label="Tienda", width=220, border_radius=8)

        with SessionLocal() as db:
            sucs = get_sucursales(db)
        e_sucursal.options = [ft.dropdown.Option(s.nombre) for s in sucs]
        e_sucursal.value   = sucursal if sucursal in [s.nombre for s in sucs] else (
            sucs[0].nombre if sucs else None)

        dlg = ft.AlertDialog(modal=True)

        def guardar(ev):
            if not all([e_nombre.value, e_apellido.value, e_dni.value, e_cargo.value]):
                self._snack("⚠️ Completa todos los campos", error=True); return
            try:
                with SessionLocal() as db:
                    actualizar_empleado(
                        db, emp_id,
                        e_nombre.value, e_apellido.value,
                        e_dni.value, e_cargo.value, e_sucursal.value,
                    )
                self.page.close(dlg)
                on_done()
                self._snack("✅ Empleado actualizado")
            except DNIExistenteError as ex:
                self._snack(f"⚠️ {ex}", error=True)
            except Exception as ex:
                self._snack(f"❌ Error: {ex}", error=True)

        dlg.title = ft.Row([
            ft.Icon(ft.Icons.EDIT_OUTLINED, color=ft.Colors.AMBER_400),
            ft.Text("Editar empleado", size=16, weight=ft.FontWeight.BOLD),
        ], spacing=10)
        dlg.content = ft.Container(
            content=ft.Column([
                ft.Row([e_nombre, e_apellido], spacing=10),
                ft.Row([e_dni, e_cargo],       spacing=10),
                e_sucursal,
            ], spacing=14, tight=True),
            width=520, padding=ft.padding.only(top=8),
        )
        dlg.actions = [
            ft.TextButton("Cancelar", on_click=lambda _: self.page.close(dlg),
                          style=ft.ButtonStyle(color=ft.Colors.GREY_400)),
            ft.ElevatedButton("💾 Guardar cambios", on_click=guardar,
                              bgcolor=ft.Colors.AMBER_700, color=ft.Colors.WHITE),
        ]
        dlg.actions_alignment = ft.MainAxisAlignment.END
        self.page.open(dlg)

    # ── Visor / descarga de QR ────────────────────────────────────────────────
    def _mostrar_qr(self, codigo: str, nombre: str):
        qr_path = get_qr_path(codigo)

        # Regenerar si no existe
        if not qr_path.exists():
            generar_qr(codigo, nombre)

        dlg = ft.AlertDialog(modal=True)

        def guardar_en_escritorio(e):
            escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
            os.makedirs(escritorio, exist_ok=True)
            destino = os.path.join(escritorio, f"QR_{codigo}.png")
            shutil.copy2(str(qr_path), destino)
            self._snack(f"✅ QR guardado en Escritorio: QR_{codigo}.png")

        dlg.title = ft.Row([
            ft.Icon(ft.Icons.QR_CODE_2, color=ft.Colors.BLUE_400),
            ft.Text(f"QR — {nombre}", size=16, weight=ft.FontWeight.BOLD),
        ], spacing=10)

        dlg.content = ft.Container(
            content=ft.Column([
                # Imagen QR
                ft.Container(
                    content=ft.Image(
                        src=str(qr_path),
                        width=240, height=240,
                        fit=ft.ImageFit.CONTAIN,
                    ),
                    bgcolor=ft.Colors.WHITE,
                    border_radius=12,
                    padding=12,
                    alignment=ft.alignment.center,
                ),
                ft.Container(height=4),
                # Código legible
                ft.Container(
                    content=ft.Column([
                        ft.Text("Código único", size=11, color=ft.Colors.GREY_500),
                        ft.Text(codigo, size=20, weight=ft.FontWeight.BOLD,
                                color=ft.Colors.BLUE_300,
                                font_family="monospace"),
                    ], spacing=2, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                    padding=ft.padding.symmetric(vertical=10),
                    alignment=ft.alignment.center,
                ),
                ft.Text(
                    "Este QR contiene el código único del empleado.\n"
                    "Imprímelo y colócalo en su credencial.",
                    size=12, color=ft.Colors.GREY_500,
                    text_align=ft.TextAlign.CENTER,
                ),
            ],
            spacing=10,
            horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            width=340,
            padding=ft.padding.only(top=8),
        )
        dlg.actions = [
            ft.TextButton("Cerrar", on_click=lambda _: self.page.close(dlg),
                          style=ft.ButtonStyle(color=ft.Colors.GREY_400)),
            ft.ElevatedButton(
                "💾 Guardar en Escritorio",
                on_click=guardar_en_escritorio,
                bgcolor=ft.Colors.BLUE_700,
                color=ft.Colors.WHITE,
                icon=ft.Icons.DOWNLOAD_OUTLINED,
            ),
        ]
        dlg.actions_alignment = ft.MainAxisAlignment.END
        self.page.open(dlg)

    # ── TAB HORARIOS ──────────────────────────────────────────────────────────
    def _tab_horarios(self):
        TURNOS_OPTS = ["Turno 1  (10:15)", "Turno 2  (14:00)",
                       "Turno 3  (16:00)", "Descanso"]
        TURNOS_VAL  = [TipoTurno.TURNO_1, TipoTurno.TURNO_2,
                       TipoTurno.TURNO_3, TipoTurno.DESCANSO]
        DIAS_LABEL  = {
            DiaSemana.LUNES:"Lunes", DiaSemana.MARTES:"Martes",
            DiaSemana.MIERCOLES:"Miércoles", DiaSemana.JUEVES:"Jueves",
            DiaSemana.VIERNES:"Viernes", DiaSemana.SABADO:"Sábado",
            DiaSemana.DOMINGO:"Domingo",
        }
        emp_drop = ft.Dropdown(label="Empleado", width=380, border_radius=8)

        def _recargar_emp_drop_h():
            with SessionLocal() as db:
                emps = get_empleados(db)
            emp_drop.options = [
                ft.dropdown.Option(f"{e.codigo_unico} - {e.nombre_completo}") for e in emps
            ]
            emp_drop.value = None

        _recargar_emp_drop_h()
        combos = {
            dia: ft.Dropdown(
                options=[ft.dropdown.Option(t) for t in TURNOS_OPTS],
                width=210, value="Descanso", border_radius=8,
            )
            for dia in list(DiaSemana)
        }
        horario_actual = ft.Column(spacing=4)

        def cargar(e=None):
            if not emp_drop.value: return
            codigo = emp_drop.value.split(" - ")[0]
            horario_actual.controls.clear()
            with SessionLocal() as db:
                emp = get_empleado_por_codigo(db, codigo)
                if not emp: return
                horarios = get_horarios_empleado(db, emp.id)
                base = {h.dia_semana: h.turno for h in horarios if h.semana_inicio is None}
                for dia, combo in combos.items():
                    turno = base.get(dia, TipoTurno.DESCANSO)
                    combo.value = TURNOS_OPTS[TURNOS_VAL.index(turno) if turno in TURNOS_VAL else 3]
                for h in horarios:
                    horario_actual.controls.append(ft.Row([
                        ft.Text(DIAS_LABEL.get(h.dia_semana, h.dia_semana.value.capitalize()),
                                width=100, size=13, color=ft.Colors.GREY_300),
                        ft.Container(
                            content=ft.Text(get_turno_nombre(h.turno), size=12,
                                            color=ft.Colors.BLUE_300),
                            padding=ft.padding.symmetric(horizontal=10, vertical=4),
                            bgcolor=ft.Colors.with_opacity(0.1, ft.Colors.BLUE_400),
                            border_radius=6,
                        ),
                    ], spacing=10))
            self.page.update()

        def guardar(e):
            if not emp_drop.value:
                self._snack("⚠️ Selecciona un empleado", error=True); return
            codigo = emp_drop.value.split(" - ")[0]
            horario = {
                dia: TURNOS_VAL[TURNOS_OPTS.index(combo.value)
                                if combo.value in TURNOS_OPTS else 3]
                for dia, combo in combos.items()
            }
            with SessionLocal() as db:
                emp = get_empleado_por_codigo(db, codigo)
                if emp: set_horario_base(db, emp.id, horario)
            cargar()
            self._snack("✅ Horario base guardado")

        emp_drop.on_change = cargar

        return ft.Row([
            ft.Container(
                content=ft.Column([
                    ft.Text("ASIGNAR HORARIO BASE", size=13,
                            weight=ft.FontWeight.BOLD, color=ft.Colors.GREY_400),
                    emp_drop,
                    ft.Container(height=4),
                    ft.Column([
                        ft.Row([
                            ft.Text(DIAS_LABEL[dia], width=110, size=13,
                                    color=ft.Colors.GREY_300),
                            combo,
                        ], spacing=10)
                        for dia, combo in combos.items()
                    ], spacing=6),
                    ft.Container(height=4),
                    ft.ElevatedButton("💾 Guardar horario base", on_click=guardar,
                                      bgcolor=ft.Colors.BLUE_700, color=ft.Colors.WHITE),
                ], spacing=12, scroll=ft.ScrollMode.AUTO),
                width=440, padding=10,
            ),
            ft.VerticalDivider(color="#21262d", width=1),
            ft.Container(
                content=ft.Column([
                    ft.Text("HORARIO ACTUAL", size=13,
                            weight=ft.FontWeight.BOLD, color=ft.Colors.GREY_400),
                    ft.Text("Selecciona un empleado para ver su horario.",
                            size=12, color=ft.Colors.GREY_600),
                    ft.Container(height=4),
                    horario_actual,
                ], spacing=8, scroll=ft.ScrollMode.AUTO),
                expand=True, padding=10,
            ),
        ], expand=True, spacing=0)

    # ── TAB ASISTENCIA ────────────────────────────────────────────────────────
    def _tab_asistencia(self):
        emp_drop = ft.Dropdown(label="Empleado", width=300, border_radius=8)

        def _recargar_emp_drop_a():
            with SessionLocal() as db:
                emps = get_empleados(db)
            emp_drop.options = [
                ft.dropdown.Option(f"{e.codigo_unico} - {e.nombre_completo}") for e in emps
            ]

        _recargar_emp_drop_a()
        desde = ft.TextField(label="Desde", width=150, border_radius=8,
                             value=(date.today()-timedelta(days=7)).isoformat())
        hasta = ft.TextField(label="Hasta", width=150, border_radius=8,
                             value=date.today().isoformat())
        tabla = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Fecha",   color=ft.Colors.GREY_400, size=12)),
                ft.DataColumn(ft.Text("Turno",   color=ft.Colors.GREY_400, size=12)),
                ft.DataColumn(ft.Text("Ingreso", color=ft.Colors.GREY_400, size=12)),
                ft.DataColumn(ft.Text("Salida",  color=ft.Colors.GREY_400, size=12)),
                ft.DataColumn(ft.Text("Estado",  color=ft.Colors.GREY_400, size=12)),
                ft.DataColumn(ft.Text("Justif.", color=ft.Colors.GREY_400, size=12)),
            ],
            rows=[],
            heading_row_color="#1c2128",
            border=ft.border.all(1, "#21262d"),
            border_radius=10,
            horizontal_lines=ft.BorderSide(1, "#21262d"),
        )

        def buscar(e=None):
            if not emp_drop.value: return
            codigo = emp_drop.value.split(" - ")[0]
            try:
                f_desde = date.fromisoformat(desde.value)
                f_hasta = date.fromisoformat(hasta.value)
            except ValueError:
                self._snack("⚠️ Formato de fecha inválido (YYYY-MM-DD)", error=True); return
            self._filas_asistencia_export.clear()
            tabla.rows.clear()
            with SessionLocal() as db:
                emp = get_empleado_por_codigo(db, codigo)
                if not emp: return
                for reg in get_registros_empleado(db, emp.id, f_desde, f_hasta):
                    estado_str = (reg.estado_asistencia.value.replace("_"," ").title()
                                  if reg.estado_asistencia else "—")
                    ingreso = reg.hora_ingreso.strftime("%H:%M") if reg.hora_ingreso else "—"
                    salida  = reg.hora_salida.strftime("%H:%M")  if reg.hora_salida  else "—"
                    turno_n = get_turno_nombre(reg.turno_asignado) if reg.turno_asignado else "—"
                    fila = [reg.fecha.isoformat(), turno_n, ingreso, salida, estado_str,
                            "Sí" if reg.ausencia_justificada else "No"]
                    self._filas_asistencia_export.append(fila)
                    tabla.rows.append(ft.DataRow(cells=[
                        ft.DataCell(ft.Text(fila[0], size=13)),
                        ft.DataCell(ft.Text(fila[1], size=13, color=ft.Colors.GREY_400)),
                        ft.DataCell(ft.Text(fila[2], size=13, color=ft.Colors.WHITE)),
                        ft.DataCell(ft.Text(fila[3], size=13, color=ft.Colors.GREY_500)),
                        ft.DataCell(ft.Text(fila[4], size=13)),
                        ft.DataCell(ft.Text("✅" if reg.ausencia_justificada else "", size=14)),
                    ]))
            self.page.update()

        return ft.Column([
            ft.Row([emp_drop, desde, hasta,
                    ft.ElevatedButton("🔍 Buscar", on_click=buscar,
                                      bgcolor=ft.Colors.BLUE_700, color=ft.Colors.WHITE)],
                   spacing=10, wrap=True),
            ft.Container(content=tabla, border_radius=10,
                         clip_behavior=ft.ClipBehavior.ANTI_ALIAS),
        ], spacing=16, scroll=ft.ScrollMode.AUTO)

    # ── TAB REPORTES ──────────────────────────────────────────────────────────
    def _tab_reportes(self):
        tabla = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Empleado",    color=ft.Colors.GREY_400, size=12)),
                ft.DataColumn(ft.Text("Cargo",       color=ft.Colors.GREY_400, size=12)),
                ft.DataColumn(ft.Text("✅ A tiempo", color=ft.Colors.GREY_400, size=12)),
                ft.DataColumn(ft.Text("🟡 T.Baja",  color=ft.Colors.GREY_400, size=12)),
                ft.DataColumn(ft.Text("🟠 T.Media", color=ft.Colors.GREY_400, size=12)),
                ft.DataColumn(ft.Text("🔴 T.Grave", color=ft.Colors.GREY_400, size=12)),
                ft.DataColumn(ft.Text("❌ Ausente", color=ft.Colors.GREY_400, size=12)),
                ft.DataColumn(ft.Text("Salud %",    color=ft.Colors.GREY_400, size=12)),
                ft.DataColumn(ft.Text("Nivel",      color=ft.Colors.GREY_400, size=12)),
            ],
            rows=[],
            heading_row_color="#1c2128",
            border=ft.border.all(1, "#21262d"),
            border_radius=10,
            horizontal_lines=ft.BorderSide(1, "#21262d"),
            column_spacing=16,
        )
        filas_export = []

        def cargar():
            tabla.rows.clear(); filas_export.clear()
            with SessionLocal() as db:
                for emp in get_empleados(db):
                    s = emp.salud
                    if not s: continue
                    nivel_str  = s.nivel_salud.value.capitalize()
                    nivel_icono = {"excelente":"🟢","regular":"🟡",
                                   "observacion":"🟠","critico":"🔴"}.get(s.nivel_salud.value,"")
                    porc = int(s.porcentaje_salud)
                    fila = [emp.nombre_completo, emp.cargo,
                            s.dias_a_tiempo, s.dias_tardanza_baja,
                            s.dias_tardanza_media, s.dias_tardanza_grave,
                            s.dias_ausente, f"{porc}%", nivel_str]
                    filas_export.append(fila)
                    tabla.rows.append(ft.DataRow(cells=[
                        ft.DataCell(ft.Text(emp.nombre_completo,
                                            weight=ft.FontWeight.W_500,
                                            color=ft.Colors.WHITE)),
                        ft.DataCell(ft.Text(emp.cargo, color=ft.Colors.GREY_400, size=13)),
                        ft.DataCell(ft.Text(str(s.dias_a_tiempo),  color=ft.Colors.GREEN_400)),
                        ft.DataCell(ft.Text(str(s.dias_tardanza_baja), color=ft.Colors.AMBER_400)),
                        ft.DataCell(ft.Text(str(s.dias_tardanza_media),color=ft.Colors.ORANGE_400)),
                        ft.DataCell(ft.Text(str(s.dias_tardanza_grave),color=ft.Colors.RED_400)),
                        ft.DataCell(ft.Text(str(s.dias_ausente),   color=ft.Colors.RED_700)),
                        ft.DataCell(ft.Text(f"{porc}%", weight=ft.FontWeight.BOLD,
                                            color=ft.Colors.WHITE)),
                        ft.DataCell(ft.Text(f"{nivel_icono} {nivel_str}")),
                    ]))
            self.page.update()

        def exportar_excel(e):
            if not filas_export:
                self._snack("⚠️ No hay datos para exportar", error=True); return
            try:
                ruta = _exportar_excel(filas_export, self._filas_asistencia_export or None)
                self._snack(f"✅ Excel guardado: {os.path.basename(ruta)}")
            except Exception as ex:
                self._snack(f"❌ Error: {ex}", error=True)

        def exportar_pdf(e):
            if not filas_export:
                self._snack("⚠️ No hay datos para exportar", error=True); return
            try:
                ruta = _exportar_pdf(filas_export, self._filas_asistencia_export or None)
                self._snack(f"✅ PDF guardado: {os.path.basename(ruta)}")
            except Exception as ex:
                self._snack(f"❌ Error: {ex}", error=True)

        cargar()

        return ft.Column([
            ft.Row([
                ft.Column([
                    ft.Text("SALUD DE RESPONSABILIDAD — MES ACTUAL", size=13,
                            weight=ft.FontWeight.BOLD, color=ft.Colors.GREY_400),
                    ft.Text("El reporte incluye el detalle de asistencia si "
                            "buscaste en la pestaña Asistencia.",
                            size=11, color=ft.Colors.GREY_600),
                ], spacing=4, expand=True),
                ft.Row([
                    ft.Container(
                        content=ft.Row([
                            ft.Icon(ft.Icons.TABLE_CHART_OUTLINED,
                                    color=ft.Colors.GREEN_400, size=16),
                            ft.Text("Exportar Excel", size=13,
                                    color=ft.Colors.GREEN_400, weight=ft.FontWeight.W_500),
                        ], spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                        on_click=exportar_excel,
                        padding=ft.padding.symmetric(horizontal=16, vertical=10),
                        bgcolor=ft.Colors.with_opacity(0.1, ft.Colors.GREEN_400),
                        border_radius=10,
                        border=ft.border.all(1, ft.Colors.with_opacity(0.2, ft.Colors.GREEN_400)),
                        ink=True,
                    ),
                    ft.Container(
                        content=ft.Row([
                            ft.Icon(ft.Icons.PICTURE_AS_PDF_OUTLINED,
                                    color=ft.Colors.RED_400, size=16),
                            ft.Text("Exportar PDF", size=13,
                                    color=ft.Colors.RED_400, weight=ft.FontWeight.W_500),
                        ], spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                        on_click=exportar_pdf,
                        padding=ft.padding.symmetric(horizontal=16, vertical=10),
                        bgcolor=ft.Colors.with_opacity(0.1, ft.Colors.RED_400),
                        border_radius=10,
                        border=ft.border.all(1, ft.Colors.with_opacity(0.2, ft.Colors.RED_400)),
                        ink=True,
                    ),
                ], spacing=10),
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
               vertical_alignment=ft.CrossAxisAlignment.START),
            ft.Container(height=4),
            ft.Container(content=tabla, border_radius=10,
                         clip_behavior=ft.ClipBehavior.ANTI_ALIAS),
        ], spacing=14, scroll=ft.ScrollMode.AUTO)

    # ── TAB SEGURIDAD ─────────────────────────────────────────────────────────
    def _tab_seguridad(self):
        user = self.page.session.get("user") or ""
        actual = ft.TextField(label="Contraseña actual", password=True, can_reveal_password=True, width=320, border_radius=8)
        nueva = ft.TextField(label="Nueva contraseña (mínimo 8)", password=True, can_reveal_password=True, width=320, border_radius=8)
        nueva2 = ft.TextField(label="Repetir nueva contraseña", password=True, can_reveal_password=True, width=320, border_radius=8)
        msg = ft.Text("", size=12)

        def cambiar(e):
            if not user:
                msg.value = "⚠️ No hay usuario admin en sesión."
                msg.color = ft.Colors.RED_400
                self.page.update()
                return
            if (nueva.value or "") != (nueva2.value or ""):
                msg.value = "⚠️ Las contraseñas nuevas no coinciden."
                msg.color = ft.Colors.AMBER_400
                self.page.update()
                return
            try:
                with SessionLocal() as db:
                    cambiar_password_admin(db, user, actual.value, nueva.value)
                actual.value = ""
                nueva.value = ""
                nueva2.value = ""
                msg.value = "✅ Contraseña actualizada."
                msg.color = ft.Colors.GREEN_400
            except Exception as ex:
                msg.value = f"❌ {ex}"
                msg.color = ft.Colors.RED_400
            self.page.update()

        def backup_db(e):
            try:
                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
                os.makedirs(escritorio, exist_ok=True)
                destino = os.path.join(escritorio, f"personal_backup_{date.today().isoformat()}.db")
                shutil.copy2(str(DB_PATH), destino)
                msg.value = f"✅ Backup guardado en Escritorio: {os.path.basename(destino)}"
                msg.color = ft.Colors.GREEN_400
            except Exception as ex:
                msg.value = f"❌ {ex}"
                msg.color = ft.Colors.RED_400
            self.page.update()

        nueva2.on_submit = cambiar

        return ft.Column(
            [
                ft.Text("SEGURIDAD", size=13, weight=ft.FontWeight.BOLD, color=ft.Colors.GREY_400),
                ft.Text("Cambia la contraseña del administrador y genera un backup de la base de datos.",
                        size=12, color=ft.Colors.GREY_500),
                ft.Divider(color="#21262d", height=20),
                ft.Text("Cambiar contraseña", size=12, weight=ft.FontWeight.BOLD, color=ft.Colors.GREY_400),
                actual,
                nueva,
                nueva2,
                ft.ElevatedButton("Actualizar contraseña", on_click=cambiar,
                                  bgcolor=ft.Colors.BLUE_700, color=ft.Colors.WHITE),
                ft.Divider(color="#21262d", height=20),
                ft.Text("Backup", size=12, weight=ft.FontWeight.BOLD, color=ft.Colors.GREY_400),
                ft.Text(f"Base de datos actual: {DB_PATH}", size=11, color=ft.Colors.GREY_600),
                ft.ElevatedButton("Guardar backup en Escritorio", on_click=backup_db,
                                  bgcolor=ft.Colors.GREEN_700, color=ft.Colors.WHITE),
                msg,
            ],
            spacing=12,
            scroll=ft.ScrollMode.AUTO,
        )
